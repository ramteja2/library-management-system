import openpyxl
from datetime import datetime
from rest_framework import status, viewsets
from rest_framework.decorators import api_view, parser_classes, action
from rest_framework.parsers import MultiPartParser, FormParser
from rest_framework.response import Response
from .models import Book
from .serializers import BookSerializer

# API viewset for books
class BookViewSet(viewsets.ModelViewSet):
    queryset = Book.objects.all()
    serializer_class = BookSerializer

    # Custom action to delete all books at once.
    @action(detail=False, methods=['delete'])
    def delete_all(self, request):
        count, _ = self.get_queryset().delete()
        return Response({"message": f"{count} books were deleted."}, status=status.HTTP_200_OK)

# File upload view for books
@api_view(['POST'])
@parser_classes([MultiPartParser, FormParser])
def upload_books(request):
    uploaded_file = request.FILES.get('file')
    if not uploaded_file:
        return Response({"error": "No file provided."}, status=status.HTTP_400_BAD_REQUEST)
    
    try:
        wb = openpyxl.load_workbook(uploaded_file)
        ws = wb.active

        created_count = 0
        # Assuming the first row is the header row.
        header = [cell.value for cell in ws[1]]
        
        # Process each row after the header.
        for row in ws.iter_rows(min_row=2, values_only=True):
            data = dict(zip(header, row))
            
            # New rule: if ACC_NUM is empty, stop processing and return an error.
            if not data.get("ACC_NUM"):
                return Response(
                    {"error": "ACC_NUM is required. Upload stopped."},
                    status=status.HTTP_400_BAD_REQUEST
                )
            
            # Convert DATE and INV_DATE fields to the desired string format if they are datetime objects.
            if data.get("DATE") and isinstance(data["DATE"], datetime):
                data["DATE"] = data["DATE"].strftime("%d-%m-%Y")
            if data.get("INV_DATE") and isinstance(data["INV_DATE"], datetime):
                data["INV_DATE"] = data["INV_DATE"].strftime("%d-%m-%Y")
            
            serializer = BookSerializer(data=data)
            if serializer.is_valid():
                serializer.save()
                created_count += 1
            else:
                # Stop processing on first error.
                return Response({"error": serializer.errors}, status=status.HTTP_400_BAD_REQUEST)

        return Response(
            {"message": f"File uploaded successfully. {created_count} books added."},
            status=status.HTTP_201_CREATED
        )
    except Exception as e:
        return Response({"error": str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
